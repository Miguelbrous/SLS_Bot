from pathlib import Path
from openpyxl import Workbook, load_workbook
from datetime import datetime
import json, re
from typing import List, Dict, Any

# ---------- utilidades base ----------
def _ensure_book(path: Path, sheet: str, headers: list):
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = sheet
        ws.append(headers)
        wb.save(path)
        return
    wb = load_workbook(path)
    if sheet not in wb.sheetnames:
        ws = wb.create_sheet(sheet)
        ws.append(headers)
        wb.save(path)

def _safe_row_date_str(v) -> str:
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    v = "" if v is None else str(v)
    return v[:10] if len(v) >= 10 else ""

def _read_sheet_dicts(path: Path, sheet: str) -> List[Dict[str, Any]]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [(h if h is not None else "") for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for r in rows[1:]:
        d = {}
        for i in range(min(len(headers), len(r))):
            d[str(headers[i])] = r[i]
        out.append(d)
    return out

# ---------- operaciones ----------
def append_operacion(excel_dir: Path, row: dict):
    path = Path(excel_dir) / "26. Plan de inversión.xlsx"
    sheet = "Operaciones"
    headers = [
        "FechaHora","Sesión","Símbolo","TF","Tipo","Riesgo(%)","Leverage",
        "Modo Tamaño","Capital abrir(€)","Nocional(USDT)","Precio entrada",
        "SL","TP1","TP2","%cerrado TP1","Precio salida","Resultado % neto",
        "Resultado € neto","Fees €","PnL bruto €","RiskScore","Confirmaciones",
        "Racha previa","Comentario","Capital cierre día(€)"
    ]
    try:
        _ensure_book(path, sheet, headers)
        wb = load_workbook(path)
        ws = wb[sheet]
        ws.append([
            row.get("FechaHora", datetime.utcnow().isoformat()),
            row.get("Sesión", ""),
            row.get("Símbolo", ""),
            row.get("TF", ""),
            row.get("Tipo", ""),
            row.get("Riesgo(%)", 0),
            row.get("Leverage", 0),
            row.get("Modo Tamaño", ""),
            row.get("Capital abrir(€)", 0),
            row.get("Nocional(USDT)", 0),
            row.get("Precio entrada", 0),
            row.get("SL", 0),
            row.get("TP1", 0),
            row.get("TP2", 0),
            row.get("%cerrado TP1", 0),
            row.get("Precio salida", 0),
            row.get("Resultado % neto", 0),
            row.get("Resultado € neto", 0),
            row.get("Fees €", 0),
            row.get("PnL bruto €", 0),
            row.get("RiskScore", 0),
            row.get("Confirmaciones", ""),
            row.get("Racha previa", ""),
            row.get("Comentario", ""),
            row.get("Capital cierre día(€)", 0),
        ])
        wb.save(path)
    except PermissionError:
        # Excel abierto; no bloqueamos el bot
        pass

def append_evento(excel_dir: Path, row: dict):
    """
    Registra eventos como COOLDOWN, RESET_DAILY, COOLDOWN_DD, CLOSE, RESUMEN_AUTOMATICO, etc.
    row["Detalle"] puede ser texto o JSON (lo tratamos en el resumen).
    """
    path = Path(excel_dir) / "26. Plan de inversión.xlsx"
    sheet = "Eventos"
    headers = ["FechaHora","Tipo","Detalle"]
    try:
        _ensure_book(path, sheet, headers)
        wb = load_workbook(path)
        ws = wb[sheet]
        ws.append([
            row.get("FechaHora", datetime.utcnow().isoformat()),
            row.get("Tipo", ""),
            row.get("Detalle", "")
        ])
        wb.save(path)
    except PermissionError:
        pass

# ---------- resumen diario ----------
def compute_resumen_diario(excel_dir: Path, date_str: str,
                           start_equity: float | None = None,
                           end_equity: float | None = None,
                           pnl_epsilon: float = 0.05) -> Dict[str, Any]:
    """
    Calcula KPIs del día 'date_str' (YYYY-MM-DD) a partir de hojas Operaciones y Eventos.
    Preferimos:
      - start_equity: override si se pasa; si no, tomamos RESET_DAILY de Eventos.
      - end_equity:   override si se pasa; si no, el último 'after' de los CLOSE del día.
    """
    path = Path(excel_dir) / "26. Plan de inversión.xlsx"
    ops = _read_sheet_dicts(path, "Operaciones")
    evs = _read_sheet_dicts(path, "Eventos")

    # start_equity desde RESET_DAILY si no llega override
    if start_equity is None:
        for e in evs:
            if _safe_row_date_str(e.get("FechaHora")) == date_str and str(e.get("Tipo")).upper() == "RESET_DAILY":
                det = str(e.get("Detalle") or "")
                det = det.replace(",", "")
                m = re.search(r"([0-9]+(?:\.[0-9]+)?)", det)
                if m:
                    start_equity = float(m.group(1))
                break
    start_equity = float(start_equity or 0.0)

    # datos de cierres
    closes = [e for e in evs
              if _safe_row_date_str(e.get("FechaHora")) == date_str
              and str(e.get("Tipo")).upper() == "CLOSE"]

    pnls = []
    afters = []
    for e in closes:
        det = str(e.get("Detalle") or "")
        try:
            d = json.loads(det)
            if "pnl" in d: pnls.append(float(d["pnl"]))
            if "after" in d: afters.append(float(d["after"]))
        except Exception:
            det2 = det.replace(",", "")
            m1 = re.search(r"pnl\s*=\s*(-?[0-9]+(?:\.[0-9]+)?)", det2)
            if m1: pnls.append(float(m1.group(1)))
            m2 = re.search(r"after\s*=\s*([0-9]+(?:\.[0-9]+)?)", det2)
            if m2: afters.append(float(m2.group(1)))

    # end_equity
    if end_equity is None:
        end_equity = afters[-1] if afters else start_equity
    end_equity = float(end_equity or 0.0)

    # trades y wins/losses
    wins = sum(1 for x in pnls if x is not None and x > pnl_epsilon)
    losses = sum(1 for x in pnls if x is not None and x < -pnl_epsilon)
    # si no hay eventos CLOSE, usamos nº de Operaciones del día
    ops_day = [o for o in ops if _safe_row_date_str(o.get("FechaHora")) == date_str]
    trades = len(closes) if closes else len(ops_day)

    # promedio de riesgo (%)
    avg_risk = round(
        sum(float(o.get("Riesgo(%)") or 0) for o in ops_day) / len(ops_day),
        4
    ) if ops_day else 0.0

    # pnl, pnl%, dd
    if start_equity > 0:
        pnl_eur = end_equity - start_equity
        pnl_pct = (pnl_eur / start_equity) * 100.0
        min_eq = min([start_equity] + [float(x) for x in afters]) if afters else start_equity
        max_dd_pct = max(0.0, (start_equity - min_eq) / start_equity * 100.0)
    else:
        pnl_eur = 0.0
        pnl_pct = 0.0
        max_dd_pct = 0.0

    winrate = round((wins / trades * 100.0), 2) if trades else 0.0

    return {
        "Fecha": date_str,
        "Start Equity": round(start_equity, 6),
        "End Equity": round(end_equity, 6),
        "PnL €": round(pnl_eur, 6),
        "PnL %": round(pnl_pct, 4),
        "Trades": trades,
        "Wins": wins,
        "Losses": losses,
        "Winrate %": winrate,
        "Avg Risk %": float(avg_risk),
        "Max DD %": round(max_dd_pct, 4),
        "Notas": ""
    }

def upsert_resumen_diario(excel_dir: Path, resumen: Dict[str, Any]):
    """
    Escribe/actualiza una fila del día en la hoja 'Resumen Diario'.
    Si ya existe la fecha, la sobreescribe; si no, la crea.
    """
    path = Path(excel_dir) / "26. Plan de inversión.xlsx"
    sheet = "Resumen Diario"
    headers = [
        "Fecha","Start Equity","End Equity","PnL €","PnL %",
        "Trades","Wins","Losses","Winrate %","Avg Risk %","Max DD %","Notas"
    ]
    try:
        _ensure_book(path, sheet, headers)
        wb = load_workbook(path)
        ws = wb[sheet]

        # buscar fila por fecha
        target = str(resumen.get("Fecha"))
        row_idx = None
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            cell = row[0].value
            cell_date = cell.strftime("%Y-%m-%d") if isinstance(cell, datetime) else (str(cell)[:10] if cell else "")
            if cell_date == target:
                row_idx = idx
                break

        values = [resumen.get(h) for h in headers]
        if row_idx:
            for j, val in enumerate(values, start=1):
                ws.cell(row=row_idx, column=j).value = val
        else:
            ws.append(values)

        wb.save(path)
    except PermissionError:
        # Excel abierto; no bloqueamos el bot
        pass
